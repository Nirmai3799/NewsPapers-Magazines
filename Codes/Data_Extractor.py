# import module
import openpyxl
import re

# load excel with its path
wrkbk = openpyxl.load_workbook("weather.xlsx")

sh = wrkbk.active

df={}
# iterate through excel and display data
i=1
while(i<sh.max_row+1):


    cell_obj = sh.cell(row=i, column=1)

    s=cell_obj.value
    s=str(s)
   

    
    
   

    s=str(s)
    s = re.sub(r'[^\w\s]', '',s)
    print(s,end=" ")
    i=i+1
    cell_obj = sh.cell(row=i, column=1)
    print(cell_obj.value,end=" ")
   

    
    

    df[s]=cell_obj.value
    i=i+1

   
       

       



print("\n")
#print(df)





   


    



print("\n")